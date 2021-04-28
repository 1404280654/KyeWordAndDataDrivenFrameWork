from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
import time
from config.VarConfig import *


class ParseExcle(object):

    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color=None) #设置字体颜色
        self.RGBDict = {'red':'FFFF3030', 'green':'FF008B00'}


    def loadWorkBook(self, excelPathAndName):
        # 将Eecel文件加载到内存，并获取workbook对象
        try:
            self.workbook = load_workbook(dataFilePath)
        except Exception as e:
            raise e
        self.excelFile = excelPathAndName
        return self.workbook

    def getSheetByName(self, sheetName):
        # 根据sheet的名获取sheet对象
        try:
            sheet = self.workbook.get_sheet_by_name(sheetName)
            return sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self, sheetIndex):
        # 根据sheet的索引号获取sheet对象
        try:
            sheetname = self.workbook.get_sheet_names()[sheetIndex]
        except Exception as e:
            raise e
        sheet = self.workbook.get_sheet_by_name(sheetname)
        return sheet

    def getRowsNumber(self, sheet):
        # 获取sheet中数据区域的结束行号
        return sheet.max_row

    def getClosNumber(self, sheet):
        # 获取sheet中数据区域的结束列号
        return sheet.max_column

    def getStartRowNumber(self, sheet):
        # 获取sheet中数据区域的开始行号
        return sheet.min_row

    def getStartColNumber(self, sheet):
        # 获取sheet中数据区域的开始列号
        return sheet.min_column

    def getRow(self, sheet, rowNo):
        # 获取sheet中的某一行，返回的是这一行所有的数据内容组成的tuple
        # 下标从1开始，sheet.rows[1]表示第一行
        try:
            self.row = []
            for i in list(sheet.rows)[rowNo-1]:
                self.row.append(i.value)
            return self.row
        except Exception as e:
            raise e

    def getColumn(self, sheet, colNo):
        # 获取sheet中的某一列，返回的是这一列所有的数据内容组成的tuple
        # 下标从1开始，sheet.rows[1]表示第一列
        try:
            self.col = []
            for i in list(sheet.columns)[colNo-1]:
                self.col.append(i.value)
            return self.col
        except Exception as e:
            raise e

    def getCellOfValue(self, sheet, coordinate = None, rowNo = None, colsNo = None):
        # 根据sheet单元格所在的位置索引获取单位格中的值，下标从1开始
        # sheet.cell(row = 1,colum = 1).value,表示Excel中的第一行第一列
        if coordinate != None:
            try:
                return sheet.cell(coordinate=coordinate).value
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colsNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell ！")


    def getCellOfObject(self, sheet, coordinate = None, rowNo = None, colsNo = None):
        # 获取某个单元格的对象，可以根据单元个所在位置的数字索引，
        # 也可以直接根据单元格的编码和坐标
        # 如getCellObject(sheet, coordinate = "A1") or
        # getCellObjiect(sheet, rowNo = 1, colNo = 2)
        if coordinate != None:
            try:
                return sheet.cell(coordinate = coordinate)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colsNo)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell ！")

    def writeCell(self, sheet, content, coordinate = None, rowNo = None, colsNo = None, style = None):
        # 根据单元格在Excel中的编码坐标或则数字所引坐标项单元格中写入数据，
        # 下标从1开始，参数style表示字体的颜色的名字，如red,yello
        if coordinate is not None:
            try:
                sheet.cell(coordinate = coordinate).value = content
                if style is not None:
                    sheet.cell(coordinate = coordinate).font = Font(color=self.RGBDict[style])
                    self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo, column=colsNo).value=content
                if style is not None:
                    sheet.cell(row=rowNo, column=colsNo).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell ！")

    def writeCellCurrentTime(self, sheet, coordinate = None, rowNo = None, colsNo = None):
        # 写入当前的时间，下标从1开始
        now = int(time.time()) # 显示为时间戳
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d%H:%M:%S", timeArray)
        if coordinate is not None:
            try:
                sheet.cell(coordinate = coordinate).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo, column=colsNo).value=currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise  e
        else:
            raise  Exception("Insufficient Coordinates of cell ！")



if __name__ == "__main__":
    pe = ParseExcle()
    pe.loadWorkBook(r"D:\\163邮箱联系人.xlsx")
    print("通过名称获取sheet对象的名字：", pe.getSheetByName(u"联系人")) # 都是为了获取对象
    print("通过index序号获取sheet的对象名子：", pe.getSheetByIndex(0))
    sheet = pe.getSheetByIndex(0)
    print("-----------   1   ------------")
    print(type(sheet))
    print(pe.getRowsNumber(sheet))
    print(pe.getClosNumber(sheet))
    print("-----------   2   ------------")
    rows = pe.getRow(sheet, 2)
    for i in rows:
        print(i)
    print("-----------   3   ------------")
    # 获取第一行第一列单元格的内容
    print(pe.getCellOfValue(sheet, rowNo=1, colsNo=1))
    pe.writeCell(sheet, u'陈俊如', rowNo=10, colsNo=10)
    pe.writeCellCurrentTime(sheet, rowNo=10, colsNo=11)
